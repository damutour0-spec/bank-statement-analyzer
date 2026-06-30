from datetime import datetime
from decimal import Decimal

import openpyxl

from statement_analyzer.exporter import export_workbook
from statement_analyzer.models import AnalysisResult, Finding, Statement, Transaction


def test_export_workbook_has_expected_sheets_and_amount_format(tmp_path):
    statement = Statement(
        file_name="sample.csv",
        file_type="csv",
        bank_name="招商银行",
        transactions=[
            Transaction(
                row_no=1,
                transaction_date=datetime(2026, 1, 1),
                summary="工资",
                counterparty_name="某公司",
                income_amount=Decimal("10000.10"),
                expense_amount=Decimal("0"),
                balance=Decimal("10000.10"),
                raw_text="2026-01-01 工资 某公司 10000.10 10000.10",
            )
        ],
    )
    analysis = AnalysisResult(
        findings=[
            Finding(
                finding_type="large_round_amount",
                severity="info",
                title="大额整数交易",
                description="测试说明",
                row_no=1,
                evidence={"amount": "10000.10"},
                suggestion="测试建议",
            )
        ],
        metrics={
            "transaction_count": 1,
            "total_income": "10000.10",
            "total_expense": "0",
            "net_flow": "10000.10",
            "monthly": {"2026-01": {"income": "10000.10", "expense": "0", "count": "1"}},
            "top_counterparties": [
                {
                    "name": "某公司",
                    "income": "10000.10",
                    "expense": "0",
                    "total_flow": "10000.10",
                    "count": "1",
                }
            ],
        },
    )
    output = tmp_path / "analysis.xlsx"

    export_workbook(statement, analysis, output)

    workbook = openpyxl.load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "报告封面",
        "标准流水",
        "异常清单",
        "汇总指标",
        "月度汇总",
        "对手方汇总",
        "规则说明",
        "原始文本",
    ]
    assert workbook["报告封面"]["A1"].value == "银行流水分析报告"
    assert workbook["标准流水"]["C1"].value == "分类"
    assert workbook["标准流水"]["C2"].value == "payroll"
    assert Decimal(str(workbook["标准流水"]["F2"].value)) == Decimal("10000.10")
    assert workbook["标准流水"]["F2"].number_format == "#,##0.00"
    assert workbook["标准流水"].freeze_panes == "A2"
    assert workbook["标准流水"].auto_filter.ref is not None
    assert workbook["异常清单"]["F1"].value == "证据"
    assert '"amount"' in workbook["异常清单"]["F2"].value
    assert workbook["月度汇总"]["D1"].value == "净流入"
    assert workbook["对手方汇总"]["E1"].value == "占比"
    assert workbook["规则说明"].max_row > 1
    assert "2026-01-01 工资" in workbook["原始文本"]["F2"].value
