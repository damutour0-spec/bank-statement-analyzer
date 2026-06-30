from __future__ import annotations

import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .models import AnalysisResult, Statement


AMOUNT_FORMAT = "#,##0.00"
PERCENT_FORMAT = "0.00%"
HEADER_FILL = "1F2937"
TITLE_FILL = "DBEAFE"

RULE_DESCRIPTIONS = [
    (
        "balance_continuity_failed",
        "余额连续性不匹配",
        "上一笔余额 + 本笔收入 - 本笔支出 与当前余额不一致。",
        "优先复核金额方向、余额、是否缺页或漏行。",
    ),
    (
        "duplicate_transaction",
        "疑似重复交易",
        "日期、摘要、金额、余额高度一致。",
        "检查是否重复导入、PDF 重复页或银行流水本身重复。",
    ),
    (
        "low_confidence",
        "低置信度交易行",
        "关键字段不完整，或来自 OCR/非结构化文本抽取。",
        "建议对照原始流水或回单人工确认。",
    ),
    (
        "sensitive_keyword",
        "敏感关键词",
        "摘要、附言或对手方命中借款、还款、贷款、网贷等关键词。",
        "结合业务背景判断是否需要说明材料。",
    ),
    (
        "large_round_amount",
        "大额整数交易",
        "交易金额较大且为整数金额。",
        "用于贷款、审计、尽调时可补充交易背景。",
    ),
    (
        "same_day_in_out",
        "当日大额进出",
        "同一天大额收入后又有较大比例支出，资金沉淀较低。",
        "复核是否为短期周转、过桥、归集或正常经营结算。",
    ),
    (
        "counterparty_concentration",
        "对手方集中度较高",
        "最大对手方流水占比过高。",
        "确认是否为关联方、主要客户、固定资金通道或正常集中结算。",
    ),
]


def export_workbook(statement: Statement, analysis: AnalysisResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()

    cover = workbook.active
    cover.title = "报告封面"
    write_cover(cover, statement, analysis)
    write_standard_transactions(workbook.create_sheet("标准流水"), statement)
    write_findings(workbook.create_sheet("异常清单"), analysis)
    write_metrics(workbook.create_sheet("汇总指标"), analysis)
    write_monthly(workbook.create_sheet("月度汇总"), analysis)
    write_counterparties(workbook.create_sheet("对手方汇总"), analysis)
    write_rule_descriptions(workbook.create_sheet("规则说明"))
    write_raw_text(workbook.create_sheet("原始文本"), statement)

    for sheet in workbook.worksheets:
        autosize(sheet)
        sheet.sheet_view.showGridLines = False
    workbook.save(path)


def write_cover(ws, statement: Statement, analysis: AnalysisResult) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = "银行流水分析报告"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("文件名", statement.file_name),
        ("文件类型", statement.file_type),
        ("识别银行", statement.bank_name),
        ("账户名称", statement.account_name),
        ("账号", statement.account_no_masked),
        ("交易笔数", len(statement.transactions)),
        ("起始日期", statement.to_summary_dict().get("start_date", "")),
        ("结束日期", statement.to_summary_dict().get("end_date", "")),
        ("整体置信度", statement.confidence),
        ("异常提示数", len(analysis.findings)),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
        ws.cell(row=row_index, column=1).font = Font(bold=True)

    severity_counts = {"high": 0, "warn": 0, "info": 0}
    for finding in analysis.findings:
        severity_counts[finding.severity] = severity_counts.get(finding.severity, 0) + 1
    start = len(rows) + 5
    ws.cell(row=start, column=1, value="异常等级汇总").font = Font(bold=True)
    ws.append(["高风险", severity_counts.get("high", 0)])
    ws.append(["警示", severity_counts.get("warn", 0)])
    ws.append(["提示", severity_counts.get("info", 0)])

    note_row = start + 5
    ws.cell(row=note_row, column=1, value="说明").font = Font(bold=True)
    ws.cell(
        row=note_row,
        column=2,
        value="本报告仅用于流水标准化、数据质量检查和辅助复核，不判断流水真伪。",
    )
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48


def write_standard_transactions(ws, statement: Statement) -> None:
    headers = [
        "行号",
        "交易时间",
        "摘要",
        "对方户名",
        "收入",
        "支出",
        "余额",
        "渠道",
        "附言",
        "置信度",
    ]
    ws.append(headers)
    style_header(ws)
    for item in statement.transactions:
        ws.append(
            [
                item.row_no,
                item.transaction_date.strftime("%Y-%m-%d %H:%M:%S") if item.transaction_date else "",
                item.summary,
                item.counterparty_name,
                item.income_amount,
                item.expense_amount,
                item.balance if item.balance is not None else "",
                item.channel,
                item.postscript,
                item.confidence,
            ]
        )
    format_amount_columns(ws, ["E", "F", "G"])
    apply_table_view(ws)


def write_findings(ws, analysis: AnalysisResult) -> None:
    ws.append(["严重度", "类型", "标题", "行号", "说明", "证据", "建议"])
    style_header(ws)
    fills = {
        "high": PatternFill("solid", fgColor="FCA5A5"),
        "warn": PatternFill("solid", fgColor="FDE68A"),
        "info": PatternFill("solid", fgColor="BFDBFE"),
    }
    for finding in analysis.findings:
        ws.append(
            [
                finding.severity,
                finding.finding_type,
                finding.title,
                finding.row_no or "",
                finding.description,
                evidence_to_text(finding.evidence),
                finding.suggestion,
            ]
        )
        fill = fills.get(finding.severity)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    wrap_columns(ws, ["E", "F", "G"])
    apply_table_view(ws)


def write_metrics(ws, analysis: AnalysisResult) -> None:
    ws.append(["指标", "值"])
    style_header(ws)
    labels = {
        "transaction_count": "交易笔数",
        "total_income": "总收入",
        "total_expense": "总支出",
        "net_flow": "净流入",
        "min_balance": "最低余额",
        "max_balance": "最高余额",
        "avg_balance": "平均余额",
    }
    amount_keys = {
        "total_income",
        "total_expense",
        "net_flow",
        "min_balance",
        "max_balance",
        "avg_balance",
    }
    for key, label in labels.items():
        value = analysis.metrics.get(key, "")
        ws.append([label, decimal_or_original(value) if key in amount_keys else value])
    format_amount_columns(ws, ["B"])
    apply_table_view(ws)


def write_monthly(ws, analysis: AnalysisResult) -> None:
    ws.append(["月份", "收入", "支出", "净流入", "交易笔数"])
    style_header(ws)
    for month, values in analysis.metrics.get("monthly", {}).items():
        income = decimal_or_zero(values.get("income"))
        expense = decimal_or_zero(values.get("expense"))
        ws.append([month, income, expense, income - expense, values.get("count")])
    format_amount_columns(ws, ["B", "C", "D"])
    apply_table_view(ws)


def write_counterparties(ws, analysis: AnalysisResult) -> None:
    ws.append(["对手方", "收入", "支出", "总流水", "占比", "交易笔数"])
    style_header(ws)
    total_flow = decimal_or_zero(analysis.metrics.get("total_income")) + decimal_or_zero(
        analysis.metrics.get("total_expense")
    )
    for item in analysis.metrics.get("top_counterparties", []):
        flow = decimal_or_zero(item.get("total_flow"))
        ratio = float(flow / total_flow) if total_flow else 0
        ws.append(
            [
                item["name"],
                decimal_or_original(item["income"]),
                decimal_or_original(item["expense"]),
                flow,
                ratio,
                item["count"],
            ]
        )
    format_amount_columns(ws, ["B", "C", "D"])
    for cell in ws["E"][1:]:
        cell.number_format = PERCENT_FORMAT
    apply_table_view(ws)


def write_rule_descriptions(ws) -> None:
    ws.append(["规则类型", "规则名称", "触发逻辑", "建议处理"])
    style_header(ws)
    for row in RULE_DESCRIPTIONS:
        ws.append(list(row))
    wrap_columns(ws, ["C", "D"])
    apply_table_view(ws)


def write_raw_text(ws, statement: Statement) -> None:
    ws.append(["行号", "交易时间", "渠道", "摘要", "原始文本"])
    style_header(ws)
    for item in statement.transactions:
        ws.append(
            [
                item.row_no,
                item.transaction_date.strftime("%Y-%m-%d %H:%M:%S") if item.transaction_date else "",
                item.channel,
                item.summary,
                item.raw_text,
            ]
        )
    wrap_columns(ws, ["E"])
    apply_table_view(ws)


def evidence_to_text(value: dict[str, Any]) -> str:
    if not value:
        return ""
    return json.dumps(value, ensure_ascii=False, default=str)


def decimal_or_original(value: Any) -> Decimal | Any:
    if value in (None, ""):
        return ""
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return value


def decimal_or_zero(value: Any) -> Decimal:
    parsed = decimal_or_original(value)
    return parsed if isinstance(parsed, Decimal) else Decimal("0")


def format_amount_columns(ws, column_letters: list[str]) -> None:
    for letter in column_letters:
        for cell in ws[letter][1:]:
            if isinstance(cell.value, Decimal):
                cell.number_format = AMOUNT_FORMAT


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")


def apply_table_view(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def wrap_columns(ws, column_letters: list[str]) -> None:
    for letter in column_letters:
        for cell in ws[letter][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
